from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy  # Importar a função copy
import os  # Importar biblioteca para verificar existência de arquivos e abrir o arquivo

def copy_sheet(source_sheet, target_sheet):
    # Copiar células e formatação
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copiar formatação
            if cell.has_style:
                new_cell.font = copy(cell.font)  # Copiar fonte
                new_cell.fill = copy(cell.fill)  # Copiar preenchimento
                new_cell.border = copy(cell.border)  # Copiar bordas
                new_cell.alignment = copy(cell.alignment)  # Copiar alinhamento
                new_cell.number_format = cell.number_format  # Copiar formato numérico

    # Copiar a mesclagem de células
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

# Caminhos dos arquivos
caminho_arquivo1 = r"C:\Users\AmandaAlves\CP SRN Dropbox\AMANDA - TI\Arquivos Amanda\FATURAMENTO DIARIO\RELATORIO PADRAO SEMANAL\01- bot rELATRIO sEMANAL - pOR lOJA.XLSX"
caminho_arquivo2 = r"C:\Users\AmandaAlves\CP SRN Dropbox\AMANDA - TI\Arquivos Amanda\FATURAMENTO DIARIO\RELATORIO PADRAO SEMANAL\02- BOT Relatrio Semanal - Por dia.xlsx"
caminho_arquivo3 = r"C:\Users\AmandaAlves\CP SRN Dropbox\AMANDA - TI\Arquivos Amanda\FATURAMENTO DIARIO\RELATORIO PADRAO SEMANAL\03- BOT Relatrio Semanal - BER - Por vendedor.xlsx"
caminho_arquivo4 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\04- BOT Relatrio Semanal - CAR - Por vendedor.xlsx"
caminho_arquivo5 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\05- BOT Relatrio Semanal - OUR - Por vendedor.xlsx"
caminho_arquivo6 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\06- BOT Relatrio Semanal - SAN - Por vendedor.xlsx"
caminho_arquivo7 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\07- BOT Relatrio Semanal - SFC - Por vendedor.xlsx"
caminho_arquivo8 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\08 - BOT Relatrio Semanal - JAC - Por vendedor.xlsx"
caminho_arquivo9 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\09 - BOT Relatrio Semanal - JAP - Por vendedor.xlsx"
caminho_arquivo10 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\10 - BOT Relatrio Semanal - EDU - Por vendedor.xlsx"
caminho_arquivo11 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\11 - BOT Relatrio Semanal - JUB - Por vendedor.xlsx"
caminho_arquivo12 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\12 - BOT Relatrio Semanal - X - Por vendedor.xlsx"
caminho_arquivo13 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\13 - BOT Relatrio Semanal - X - Por vendedor.xlsx"
caminho_arquivo14 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\14 - BOT Relatrio Semanal - PLS - Por vendedor.xlsx"
caminho_arquivo15 = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\RELATORIO PADRAO SEMANAL\15 - BOT Relatrio Semanal - MOC - Por vendedor.xlsx"

caminho_saida = r"C:\Users\Amanda\CP SRN Dropbox\AMANDA - TI\Arquivos Amanda\POWER BI\faturamento diario\FATURAMENTO DIARIO\PADRÃO RELATORIOS - PYTHON\SAIDA.xlsx"

# Lista de arquivos para carregar
caminhos_arquivos = [caminho_arquivo1, caminho_arquivo2, caminho_arquivo3, caminho_arquivo4, caminho_arquivo5, caminho_arquivo6, caminho_arquivo7, caminho_arquivo8, caminho_arquivo9, caminho_arquivo10, caminho_arquivo11, caminho_arquivo12, caminho_arquivo13, caminho_arquivo14,caminho_arquivo15]

# Criar uma nova pasta de trabalho
wb_saida = Workbook()

# Remover a aba padrão criada pelo Workbook
if "Sheet" in wb_saida.sheetnames:
    std = wb_saida["Sheet"]
    wb_saida.remove(std)

# Carregar as planilhas e copiar para wb_saida
for caminho_arquivo in caminhos_arquivos:
    try:
        if os.path.exists(caminho_arquivo):
            wb = load_workbook(caminho_arquivo)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                # Nomear a aba de destino com o índice do arquivo para evitar duplicação
                ws_saida = wb_saida.create_sheet(title=f'{sheet}_{caminhos_arquivos.index(caminho_arquivo) + 1}')
                copy_sheet(ws, ws_saida)
        else:
            print(f"Arquivo não encontrado: {caminho_arquivo}")
    except Exception as e:
        print(f"Erro ao processar {caminho_arquivo}: {e}")

# Salvar a nova pasta de trabalho
try:
    wb_saida.save(caminho_saida)
    print("Planilhas unidas e salvas com sucesso!")
    
    # Abrir o arquivo automaticamente após salvar
    os.startfile(caminho_saida)
    print("Arquivo de saída aberto com sucesso!")
except Exception as e:
    print(f"Erro ao salvar ou abrir o arquivo de saída: {e}")


